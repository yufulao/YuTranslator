from ui.batch.model import BatchModel
from ui.batch.view import BatchView
import os
import shutil
from openpyxl import load_workbook


class BatchCtrl:
    def __init__(self):
        self._model = BatchModel()
        self._view = BatchView()
        self._bind_event()
        # print(self._model.model_dict)

    #打开界面
    def open(self):
        return self._view.open()
    
    def _bind_event(self):
        self._view.action_translate = self._translate
        
    def _parse_ranges(self, ranges_str):
        # 将 "C4:C5, C8:C58" -> ["C4:C5", "C8:C58"]
        return [r.strip() for r in ranges_str.split(",") if r.strip()]

    def _translate(self, input_path, sheet_name, input_range_str, output_range_str, overwrite, output_dir):
        try:
            if not input_path or not sheet_name or not input_range_str or not output_range_str or (not overwrite and not output_dir):
                return "❌ 请填写所有字段"

            input_ranges = self._parse_ranges(input_range_str)
            output_ranges = self._parse_ranges(output_range_str)

            if len(input_ranges) != len(output_ranges):
                return "❌ 输入范围和输出范围数量不匹配"

            file_name = os.path.basename(input_path)
            output_path = os.path.join(output_dir, "output_"+file_name)
            if not overwrite:
                shutil.copyfile(input_path, output_path)
                working_path = output_path
            else:
                working_path = input_path

            wb = load_workbook(working_path)
            if sheet_name not in wb.sheetnames:
                return f"❌ 找不到工作表：{sheet_name}"

            sheet = wb[sheet_name]
            
            
            from ui.main_controller import MainController
            home_ctrl = MainController().instance.home_ctrl
            state = MainController().instance.state
            if "home" not in state:
                return "未设置翻译界面参数"
            home_info = state["home"]
            target_language = home_info["target_language"]
            model_name = home_info["model_name"]
            lora_input = home_info["lora_input"]
            temperature = home_info["temperature"]
            max_tokens = home_info["max_tokens"]
            top_p = home_info["top_p"]
            arg = (temperature, max_tokens, top_p)
            
            translator = home_ctrl.get_translator(model_name, lora_input)
            
            
            state = MainController().instance.state
            if "home" not in state:
                return "未设置翻译界面参数"
            asdsa = state["home"]
            print(asdsa["target_language"])
            
            
            for input_range, output_range in zip(input_ranges, output_ranges):
                input_cells = list(sheet[input_range])
                output_cells = list(sheet[output_range])
                if len(input_cells) != len(output_cells):
                    return f"❌ 区块长度不一致: {input_range} vs {output_range}"

                for i in range(len(input_cells)):
                    in_cell = input_cells[i][0]
                    out_cell = output_cells[i][0]
                    if in_cell.value is not None:
                        input_text = str(in_cell.value)
                        out_cell.value = home_ctrl.translate_by_translator(translator, input_text, target_language, *arg)

            wb.save(working_path)
            return f"✅ 翻译完成，文件保存到：{working_path}"
        except Exception as e:
            return f"❌ 错误: {str(e)}"

